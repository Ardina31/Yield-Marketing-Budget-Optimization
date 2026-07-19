"""Generates downloadable PDF, Excel, and CSV performance reports."""
import csv
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

TEAL = colors.HexColor("#0F766E")
INK = colors.HexColor("#0B1220")
SLATE = colors.HexColor("#64748B")
EMBER = colors.HexColor("#E8543F")
LIGHT_ROW = colors.HexColor("#F1F5F9")


def generate_pdf_report(user_name: str, kpis: dict, rows: list[dict]) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleStyle", parent=styles["Title"], textColor=INK, fontSize=22, spaceAfter=2,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"], textColor=SLATE, fontSize=10, spaceAfter=18,
    )
    section_style = ParagraphStyle(
        "Section", parent=styles["Heading2"], textColor=INK, fontSize=13, spaceBefore=18, spaceAfter=8,
    )

    story = [
        Paragraph("Marketing Performance Report", title_style),
        Paragraph(
            f"Prepared for {user_name} &nbsp;•&nbsp; Generated {datetime.now().strftime('%B %d, %Y at %I:%M %p')}",
            subtitle_style,
        ),
    ]

    kpi_data = [
        ["Total Budget", "Total Revenue", "Net Profit", "ROI"],
        [
            f"${kpis['total_budget']:,.0f}",
            f"${kpis['total_revenue']:,.0f}",
            f"${kpis['total_profit']:,.0f}",
            f"{kpis['roi']:.1f}%",
        ],
    ]
    kpi_table = Table(kpi_data, colWidths=[1.6 * inch] * 4)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), INK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, 1), 14),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 1), (-1, 1), TEAL),
        ("BACKGROUND", (0, 1), (-1, 1), LIGHT_ROW),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
    ]))
    story.append(kpi_table)

    story.append(Paragraph("Campaign Performance", section_style))
    table_header = ["Campaign", "Channel", "Cost", "Revenue", "ROI", "Conversions"]
    table_data = [table_header]
    for r in sorted(rows, key=lambda x: x["roi"], reverse=True):
        table_data.append([
            r["campaign"].name[:28],
            r["campaign"].channel.name,
            f"${r['cost']:,.0f}",
            f"${r['revenue']:,.0f}",
            f"{r['roi']:.1f}%",
            f"{r['conversions']:,}",
        ])

    perf_table = Table(table_data, colWidths=[1.7 * inch, 1.2 * inch, 0.9 * inch, 0.9 * inch, 0.7 * inch, 0.9 * inch])
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), TEAL),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
    ]
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), LIGHT_ROW))
    perf_table.setStyle(TableStyle(style_cmds))
    story.append(perf_table)

    if kpis.get("best_channel"):
        story.append(Paragraph("Insights", section_style))
        story.append(Paragraph(
            f"<b>{kpis['best_channel']}</b> is the strongest performing channel by net profit. "
            f"Blended ROI across all campaigns is <b>{kpis['roi']:.1f}%</b> with a conversion rate of "
            f"<b>{kpis['conversion_rate']:.1f}%</b>.",
            styles["Normal"],
        ))

    story.append(Spacer(1, 24))
    story.append(Paragraph(
        "Generated automatically by Yield — Marketing Budget Optimization Platform.",
        ParagraphStyle("Footer", parent=styles["Normal"], textColor=SLATE, fontSize=8),
    ))

    doc.build(story)
    buf.seek(0)
    return buf


def generate_excel_report(kpis: dict, rows: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Campaign Performance"

    header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["Campaign", "Channel", "Status", "Cost", "Revenue", "Profit", "ROI %", "Conversions", "CTR %", "CPA"]

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in sorted(rows, key=lambda x: x["roi"], reverse=True):
        ws.append([
            r["campaign"].name,
            r["campaign"].channel.name,
            r["campaign"].status,
            round(r["cost"], 2),
            round(r["revenue"], 2),
            round(r["revenue"] - r["cost"], 2),
            round(r["roi"], 1),
            r["conversions"],
            round(r["ctr"], 2),
            round(r["cpa"], 2),
        ])

    for i, width in enumerate([26, 16, 12, 12, 12, 12, 10, 12, 10, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    if len(rows) > 0:
        chart = BarChart()
        chart.title = "Cost vs Revenue by Campaign"
        chart.y_axis.title = "Amount ($)"
        chart.style = 10
        data = Reference(ws, min_col=4, max_col=5, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 22
        chart.height = 10
        ws.add_chart(chart, f"A{ws.max_row + 3}")

    summary = wb.create_sheet("Summary")
    summary_rows = [
        ("Total Campaigns", kpis["total_campaigns"]),
        ("Total Budget", round(kpis["total_budget"], 2)),
        ("Total Revenue", round(kpis["total_revenue"], 2)),
        ("Net Profit", round(kpis["total_profit"], 2)),
        ("Blended ROI %", round(kpis["roi"], 2)),
        ("Conversion Rate %", round(kpis["conversion_rate"], 2)),
        ("CTR %", round(kpis["ctr"], 2)),
        ("CPC", round(kpis["cpc"], 2)),
        ("CPA", round(kpis["cpa"], 2)),
        ("Best Channel", kpis.get("best_channel") or "-"),
    ]
    summary.append(["Metric", "Value"])
    for col in (1, 2):
        c = summary.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
    for label, value in summary_rows:
        summary.append([label, value])
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_csv_report(rows: list[dict]) -> io.StringIO:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Campaign", "Channel", "Status", "Cost", "Revenue", "Profit", "ROI %", "Conversions", "CTR %", "CPA"])
    for r in sorted(rows, key=lambda x: x["roi"], reverse=True):
        writer.writerow([
            r["campaign"].name,
            r["campaign"].channel.name,
            r["campaign"].status,
            round(r["cost"], 2),
            round(r["revenue"], 2),
            round(r["revenue"] - r["cost"], 2),
            round(r["roi"], 1),
            r["conversions"],
            round(r["ctr"], 2),
            round(r["cpa"], 2),
        ])
    buf.seek(0)
    return buf
